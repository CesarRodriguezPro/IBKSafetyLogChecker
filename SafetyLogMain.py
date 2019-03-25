# !/usr/bin/env python3
# -*- coding: utf8 -*-

import platform, subprocess, os, csv
import openpyxl
import requests
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from gsheets import Sheets
from openpyxl.styles import Alignment, Font, Border, Side


###################################### settings app ########################################################
DIR_PATH = os.path.abspath(os.path.dirname(__file__))
with open("Google_url.txt", "r") as file_to_open:
    GOOGLE_URL = file_to_open.read() 
with open('TimeStation_Key.txt', "r") as file_open:
    API_KEY = file_open.read()
CURRENT_EMPLOYEES_DATA = 'CURRENT_EMPLOYEES_DATA.txt'
G_BEFORE_PD = "G_BEFORE_PD.csv"
sheets = Sheets.from_files('credentials.json')
CODE = 37
GOOGLE_DATA = 'google_data.xlsx'
GOOGLE_PATH = os.path.join(DIR_PATH, GOOGLE_DATA)
############################################################################################################


today = datetime.today().strftime('%Y-%m-%d %H:%M')
today_year = datetime.today().strftime('%Y')
current_os = platform.system()
clean_screen = os.system('cls') if current_os == "Windows" else os.system('clear')


def open_file(path):
    if platform.system() == "Windows":
        os.startfile(path)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


def greetings():
    global locations
    locations = input("Please introduce your location \n\n --->  ")
    active = CreatedReport()
    active.run()


class GettingDataForReport:

    def __init__(self):
        pass
        # self.google_sheet_process()
        # self.get_data(CURRENT_EMPLOYEES_DATA)

    def google_sheet_process(self):

        s = sheets.get(GOOGLE_URL)
        s.sheets[0].to_csv(G_BEFORE_PD, encoding='utf-8', dialect='excel')
        google_data = pd.read_csv(G_BEFORE_PD)
        google_data["Name"] = google_data['Name'].apply(lambda x: x.lower())
        google_data.to_excel(GOOGLE_DATA)

    def time_return(self, x):

        def convert_time(x):
            v = datetime.strptime(x, '%Y-%m-%d')
            return v
        try:
            years_ago = x - relativedelta(years=-5)
            return years_ago.strftime('%Y-%m-%d')
        except:
            conve = convert_time(x)
            years_ago = conve - relativedelta(years=-5)
            return years_ago.strftime('%Y-%m-%d')

    def get_data(self, out_file, code=37):
        #  this is preconfigure for current employee status  for the whole company
        #  by we can use any other code that not require a date input.

        try:
            ship_api_url = "https://api.mytimestation.com/v0.1/reports/?api_key={}&id={}&exportformat=csv".format(API_KEY, CODE)
            url_responds = requests.get(ship_api_url)
            CURRENT_EMPLOYEES_DATA = url_responds.text
            save_text = open(out_file, 'w', encoding='utf8')
            save_text.write(CURRENT_EMPLOYEES_DATA)
            save_text.close()
            print('--connection made')
        except:
            print('there was a problem downloading the information from the server')
            print('please check the internet connection.')
            input('')

    def orginazing_timesation_data(self):
        """ this get the information from timestation current and give back 
        a list of names with devices"""

        item_to_return = {}
        with open(CURRENT_EMPLOYEES_DATA, 'r', newline="") as file_to_open:
            reader = csv.DictReader(file_to_open, delimiter=',')
            for row in reader:
                if row['Status'] == 'In' and locations in row['Current Department']:
                    item_to_return.setdefault(row["Device"], []).append([row['Name'].lower(), row['Current Department']])

        print('---sort finish')
        return item_to_return

    def read_info_google(self):

        file_excel = openpyxl.load_workbook(GOOGLE_PATH)
        sheet_names = file_excel.sheetnames
        sheet_name = sheet_names[0]
        active_sheet = file_excel[sheet_name]
        rows_number = active_sheet.max_row

        number = 3
        db = {}
        for x in range(rows_number-2):
            name = active_sheet['c{}'.format(number)].value
            license_number = active_sheet['d{}'.format(number)].value
            issued_date = active_sheet['e{}'.format(number)].value
            expiration_date = active_sheet['f{}'.format(number)].value
            db[name.lower()] = [license_number, issued_date, expiration_date]
            number += 1

        return db

    def run(self):

        dict_employees = self.orginazing_timesation_data()
        db = self.read_info_google()

        dict_names = [items for items in dict_employees.values()]

        dict_location = {}
        for items in dict_names:
            for item in items:
                dict_location[item[0].lower()] = item[1]

        dict_full_data = {}
        for device, names in dict_employees.items():
            for name in names:
                if name[0] not in db.keys():
                    no_osha = {}
                    no_osha[name[0]] = ["none", "None", "None",name[1]]
                    dict_full_data.setdefault(device, []).append(no_osha)

        for g_name, g_info in db.items():
            for device, list_names in dict_employees.items():
                if g_name in [name[0] for name in list_names]:  
                    g_dict = {}
                    g_info.append(dict_location[g_name])
                    g_dict[g_name] = g_info
                    dict_full_data.setdefault(device, []).append(g_dict)
        
        return dict_full_data


class CreatedReport:

    def __init__(self):

        self.erase_folders(locations)
        self.active = GettingDataForReport()
        self.dict_full_data = self.active.run()

    def erase_folders(self, folder):

        folder_to_empty = os.path.join(DIR_PATH, folder)
        if os.path.isdir(folder_to_empty):
            filelist = [f for f in os.listdir(folder_to_empty)]
            for item in filelist:
                os.remove(os.path.join(folder_to_empty, item))

    def display_info(self):
        ''' is created to see the information recollected without creating the sheets'''

        number = 1
        for device_name, list_items in self.dict_full_data.items():
            for dict_with_names in list_items:
                for names, values in dict_with_names.items():
                    print(f"{number:5} - {device_name:26} -- {names:25} -- {values}")
                    number += 1

    def full_border(self):
        border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        return border_style

    def heather(self, ws, device, attendees=0):

        ws.merge_cells("A1:E1")
        ws.merge_cells("C2:D2")
        ws.merge_cells("C3:D3")
        ws.merge_cells("C4:D4")

        ws['A1'] = "PRE-SHIFT SAFETY MEETING"
        ws['A2'] = "Project"
        ws['A3'] = "Contractor"
        ws['A4'] = "TRADE"
        ws['C2'] = "Date/Time"
        ws['C3'] = "Number of Attendees"
        ws['C4'] = "Foreman:"

        ws['B2'] = f'{self.getting_current_location()}'.title()
        ws['E2'] = datetime.today().strftime('%m/%d/%Y %I:%M %p')
        ws['E3'] = attendees
        ws['E4'] = f'{device}'.title()   # Foreman Name

        #### styles ###
        ws['A1'].font = Font(size=22, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        rd = ws.row_dimensions[1]
        rd.height = 25

        ws['B2'].font = Font(bold=True)
        ws['E3'].font = Font(size=14)
        ws['E4'].font = Font(size=12, bold=True)
        for x in range(2, 4):
            ws[f'A{x}'].font = Font(size=10)
            ws[f'C{x}'].font = Font(size=10)

        ws['E1'].border = self.full_border()
        ws['A1'].border = self.full_border()
        ws['A2'].border = self.full_border()
        ws['A3'].border = self.full_border()
        ws['A4'].border = self.full_border()
        ws['B3'].border = self.full_border()
        ws['B4'].border = self.full_border()
        ws['C2'].border = self.full_border()
        ws['C3'].border = self.full_border()
        ws['C4'].border = self.full_border()
        ws['B2'].border = self.full_border()
        ws['E2'].border = self.full_border()
        ws['E3'].border = self.full_border()
        ws['E4'].border = self.full_border()
        ws['B2'].alignment = Alignment(horizontal='center')
        ws['E2'].alignment = Alignment(horizontal='center')
        ws['E3'].alignment = Alignment(horizontal='center')
        ws['E4'].alignment = Alignment(horizontal='center')

    def body(self, ws, employees_list):
        # --------------------------------- right column  ---------------------------------------------------
        ws.merge_cells("A5:B8")
        ws['A5'] = '''SAFETY DISCUSSION
         (REVIEW ACTIVITIES/TASK TO BE PERFORMED INCLUDING SAFETY CONCERNS OR RISK WITH WORK)'''
        ws['A5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws['A5'].border = self.full_border()
        for x in range(9, 42):
            ws.merge_cells(f"A{x}:B{x}")
            ws[f'A{x}'].border = self.full_border()

        # ------------------------------------ left column  ----------------------------------------------------
        # style and border
        ws.merge_cells('C5:E5')
        ws['C5'] = 'List of Attendees'
        ws['C5'].alignment = Alignment(horizontal='center')
        ws['C5'].border = self.full_border()
        ws['E5'].border = self.full_border()

        number = 1
        for y in range(6, 42):
            ws.merge_cells(f'D{y}:E{y}')
            ws[f'D{y}'].border = self.full_border()
            ws[f'C{y}'].border = self.full_border()
            ws[f'E{y}'].border = self.full_border()
            ws[f'C{y}'] = number
            number += 1

        for row, items in enumerate(employees_list, 6):
            for name, values in items.items():
                ws[f'D{row}'] = name.title()

        count = len(employees_list)
        return count

    def footer(self, attendees, ws):

        if attendees > 36:
            start_row = 42 + (attendees - 36)
        else:
            start_row = 42

        for y in range(start_row, (start_row + 5)):
            ws.merge_cells(f'A{y}:B{y}')
            ws.merge_cells(f'C{y}:E{y}')

            ws[f'A{y}'].alignment = Alignment(horizontal='center')
            ws[f'C{y}'].alignment = Alignment(horizontal='center')
            ws[f'A{y}'].border = self.full_border()
            ws[f'C{y}'].border = self.full_border()
            ws[f'E{y}'].border = self.full_border()

        ws[f'A{start_row}'] = 'WORKSIDE HAZARD'
        ws[f'C{start_row}'] = 'PLAN TO ELIMINATE'
        ws[f'A{start_row + 2}'] ='ADDITIONAL COMMENTS'
        ws[f'A{start_row + 4}'] = 'COMPETENT PERSON CONDUCTING MEETING NAME/SIGNATURE'
        ws[f'A{start_row + 4}'].alignment = Alignment(wrap_text=True, horizontal='center')
        ws.row_dimensions[start_row + 4].height = 28

    def general_style(self, ws):

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 20

        for y in range(6, 42):  # this is to format the body rows
            ws.row_dimensions[y].height = 13
            ws[f'C{y}'].font = Font(size=10)
            ws[f'D{y}'].font = Font(size=10)
            ws[f'A{y-1}'].font = Font(size=10)

    def getting_current_location(self):
        with open(CURRENT_EMPLOYEES_DATA, 'r', newline="") as file_to_open:
            reader = csv.DictReader(file_to_open, delimiter=',')
            projects = [row['Current Department']for row in reader if locations in row['Current Department']]
            return(projects[0])

    def run(self):

        for device, employees_list in self.dict_full_data.items():
            None if os.path.isdir(locations) else os.mkdir(locations)
            wb = openpyxl.Workbook()
            ws = wb.active
            name_for_out = "{}".format(os.path.join(DIR_PATH, os.path.join(f"{locations}", f'{device}.xlsx')))

            count = self.body(ws=ws, employees_list=employees_list)
            self.heather(ws=ws, device=device, attendees=count)
            self.footer(attendees= count, ws=ws)
            self.general_style(ws=ws)

            wb.save(name_for_out)
            wb.close()

        open_file(locations)


if __name__ == '__main__':
    print("____________ Welcome ____________")
    greetings()

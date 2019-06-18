
import os, platform, datetime, subprocess
import pandas as pd 
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from Email_preparer import send_email


######################## settings ##########################
with open('TimeStation_Key.txt', 'r') as open_file:
    API_KEY = open_file.read()
CODE = 37
URL = 'https://api.mytimestation.com/v0.1/reports/?api_key={}&id={}&exportformat=csv'.format(API_KEY, CODE)
today = datetime.datetime.today()
LIST_LOCATIONS = '262 511 161 199'.split()
############################################################


###################### Multi os Support ####################
DIR_PATH = os.path.abspath(os.path.dirname(__file__))
current_os = platform.system()
clean_screen = os.system('cls') if current_os == "Windows" else os.system('clear')

def open_file(path):
    if platform.system() == "Windows":
        os.startfile(path)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])
###########################################################


class GettingDataForReport:
    ''' this gets the informatin for timestation Server using pandas to get request 
    - gets =  data from timestation 
    - returns = pandas dataframe with names filtered by location and Status "IN" '''

    def __init__(self):
        self.URL = URL
        self.raw_current = pd.read_csv(self.URL)
        self.current_in = self.raw_current[self.raw_current['Status'].str.contains('In')]
    
    def run(self, location): 
        filter_data = self.current_in[self.current_in['Current Department'].str.contains(location)]
        return filter_data


class CreatedReport:
    ''' creates the report for according to OSHA standards
    - gets: Pandas dataframe, location
    - returns: Nothing '''
    
    def __init__(self, data, location):
        self.location = location
        self.erase_folders(self.location)
        self.data = data.to_dict('index')

    def erase_folders(self, folder):
        ''' this ensure that data inside the folder are erase before the program add new files
        - get: folder name - string '''

        folder_to_empty = os.path.join(DIR_PATH, folder)
        if os.path.isdir(folder_to_empty):
            filelist = [f for f in os.listdir(folder_to_empty)]
            for item in filelist:
                os.remove(os.path.join(folder_to_empty, item))

    def full_border(self):
        ''' style - full boders for excel sheet by square'''

        border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                              top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        return border_style

    def heather(self, ws, device, location, attendees=0):
        ''' format the heather of the page that will be created 
        -gets: Device(foreman name)-string, location-string, attendees-int '''
        
        ws.merge_cells("A1:E1")
        ws.merge_cells("C2:D2")
        ws.merge_cells("C3:D3")
        ws.merge_cells("C4:D4")

        ws['A1'] = "PRE-SHIFT SAFETY MEETING"
        ws['A2'] = "Project"
        ws['A3'] = "Contractor"
        ws['A4'] = "TRADE"
        ws['B4'] = 'IBK Construction Group'
        ws['C2'] = "Date/Time"
        ws['C3'] = "Number of Attendees"
        ws['C4'] = "Foreman:"

        ws['B2'] = f'{location}'.title()
        ws['E2'] = f"{datetime.datetime.today().strftime('%m/%d/%Y')} 7:00 AM"
        ws['E3'] = attendees
        ws['E4'] = f'{device}'.title()   # Foreman Name

        #### styles ###
        ws['A1'].font = Font(size=22, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        rd = ws.row_dimensions[1]
        rd.height = 25

        ws['B2'].font = Font(bold=True)
        ws['E3'].font = Font(size=14)
        ws['E4'].font = Font(size=12, bold=True)
        for x in range(2, 4):
            ws[f'A{x}'].font = Font(size=10)
            ws[f'C{x}'].font = Font(size=10)

        ws['E1'].border = self.full_border()
        ws['A1'].border = self.full_border()
        ws['A2'].border = self.full_border()
        ws['A3'].border = self.full_border()
        ws['A4'].border = self.full_border()
        ws['B3'].border = self.full_border()
        ws['B4'].border = self.full_border()
        ws['C2'].border = self.full_border()
        ws['C3'].border = self.full_border()
        ws['C4'].border = self.full_border()
        ws['B2'].border = self.full_border()
        ws['E2'].border = self.full_border()
        ws['E3'].border = self.full_border()
        ws['E4'].border = self.full_border()
        ws['B2'].alignment = Alignment(horizontal='center')
        ws['E2'].alignment = Alignment(horizontal='center')
        ws['E3'].alignment = Alignment(horizontal='center')
        ws['E4'].alignment = Alignment(horizontal='center')

    def body(self, ws, employees_list):
        ''' format and edit the body of the page
        -gets: ws-openpyxl page activated, employees_list -  list of names of employees
        -return: count - int, name_location[0] - string '''

        total_entrys = 42 if len(employees_list) < 36 else (len(employees_list)-36) + 42

        # --------------------------------- right column  ---------------------------------------------------
        ws.merge_cells("A5:B8")
        ws['A5'] = '''SAFETY DISCUSSION
         (REVIEW ACTIVITIES/TASK TO BE PERFORMED INCLUDING SAFETY CONCERNS OR RISK WITH WORK)'''
        ws['A5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws['A5'].border = self.full_border()
        
        square = '\u25A2'
    
        ws['A9']  = square+ ' Fall Protection:'
        ws['A11'] = square+ ' Control Access Zone:'
        ws['A13'] = square+ ' Proper use of pneumatic tools'
        ws['A15'] = square+ ' Safety of Handset Forms installation.'
        ws['A17'] = square+ ' Safety of Plywood installation.'
        ws['A19'] = square+ ' Ergonomics'
        ws['A21'] = square+ ' hazardous Materials'
        ws['A23'] = square+ ' OSHA Silica Standard'
        ws['A25'] = square+ ' Use of Proper PPE:'
        ws['A27'] = square+ ' Proper use of Electrical Tools'
        ws['A29'] = square+ ' Safety of TITAN Installation.'
        ws['A31'] = square+ ' Safety of Rebar installation.'
        ws['A33'] = square+ ' Safety of Hosting and Lifting.'
        ws['A35'] = square+ ' Unusual Weather Conditions.'
        


        for x in range(9, total_entrys):
            ws.merge_cells(f"A{x}:B{x}")
            ws[f'A{x}'].border = self.full_border()

        # ------------------------------------ left column  ----------------------------------------------------
        ws.merge_cells('C5:E5')
        ws['C5'] = 'List of Attendees'
        ws['C5'].alignment = Alignment(horizontal='center')
        ws['C5'].border = self.full_border()
        ws['E5'].border = self.full_border()

        number = 1
        for y in range(6, total_entrys):
            ws.merge_cells(f'D{y}:E{y}')
            ws[f'D{y}'].border = self.full_border()
            ws[f'C{y}'].border = self.full_border()
            ws[f'E{y}'].border = self.full_border()
            ws[f'C{y}'] = number
            number += 1

        for row, items in enumerate(employees_list, 6):
            ws[f'D{row}'] = items[0].title()
        # ---------------------------------------------------------------------------------------------------------

        count = len(employees_list)
        name_location = [x[1] for x in employees_list]
        return count, name_location[0]

    def footer(self, attendees, ws):

        if attendees > 36:
            start_row = 42 + (attendees - 36)
        else:
            start_row = 42

        for y in range(start_row, (start_row + 5)):
            ws.merge_cells(f'A{y}:B{y}')
            ws.merge_cells(f'C{y}:E{y}')

            ws[f'A{y}'].alignment = Alignment(horizontal='center')
            ws[f'C{y}'].alignment = Alignment(horizontal='center')
            ws[f'A{y}'].border = self.full_border()
            ws[f'C{y}'].border = self.full_border()
            ws[f'E{y}'].border = self.full_border()

        ws[f'A{start_row}'] = 'WORKSIDE HAZARD'
        ws[f'C{start_row}'] = 'PLAN TO ELIMINATE'
        ws[f'A{start_row + 2}'] ='ADDITIONAL COMMENTS'
        ws[f'A{start_row + 4}'] = 'COMPETENT PERSON CONDUCTING MEETING NAME/SIGNATURE'
        ws[f'A{start_row + 4}'].alignment = Alignment(wrap_text=True, horizontal='center')
        ws.row_dimensions[start_row + 4].height = 28

    def general_style(self, ws):

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 30

        for y in range(6, 42):  # this is to format the body rows
            ws.row_dimensions[y].height = 13
            ws[f'C{y}'].font = Font(size=10)
            ws[f'D{y}'].font = Font(size=10)
            ws[f'A{y-1}'].font = Font(size=10)

    def convert_path_to_linux(self, device):
        ''' for linux system, saving files with no space is better.
        this clean the name of files before there are save.'''

        device = device.strip()
        device = device.replace(' ', '_')
        device = device.replace(',', "")
        return device

    def run(self):

        None if os.path.isdir(self.location) else os.mkdir(self.location)
        raw_data = [item for item in self.data.values()]
        dict_data = {}
        for x in raw_data:
            dict_data.setdefault(x['Device'],[]).append([x['Name'], x['Current Department']])
        
        for device, list_of_employees in dict_data.items():
            wb = openpyxl.Workbook()
            ws = wb.active
            name_for_out = "{}".format(os.path.join(DIR_PATH, os.path.join(f"{self.location}", f'{self.convert_path_to_linux(str(device))}.xlsx')))
            count, name_location = self.body(ws=ws, employees_list=list_of_employees)
            self.heather(ws=ws, device=device, attendees=count, location=name_location)
            self.footer(attendees= count, ws=ws)
            self.general_style(ws=ws)
            wb.save(name_for_out)
            wb.close()


def convert_to_pdf(location):
    file_dir = os.path.abspath(os.path.join(DIR_PATH, location))
    list_file_names = os.listdir(file_dir)
    for x in list_file_names:
        fullname = os.path.abspath(os.path.join(file_dir, x))
        print(fullname)
        os.system(f'"C:\Program Files\LibreOffice\program\soffice.bin"  --convert-to pdf {fullname} --outdir {file_dir}')


if __name__ == "__main__":

    getting_data = GettingDataForReport()
    for location in LIST_LOCATIONS:
        data = getting_data.run(location=location)
        active = CreatedReport(data=data, location=location)
        active.run()
        convert_to_pdf(location=location)
        send_email(location)
